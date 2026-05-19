from __future__ import annotations

from pathlib import Path
from shutil import copy2
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from config.constants import (
    BP_SHEET_NAME,
    INPUT_WORKBOOK_PATH,
    MASTER_SHEET_NAME,
    OUTPUT_DIRECTORY,
    REQD_MC_SHEET_NAME,
    WORKING_DIRECTORY,
    WORKING_WORKBOOK_FILENAME,
)


def resolve_input_workbook_path() -> Path:
    return Path(INPUT_WORKBOOK_PATH)


def resolve_working_workbook_path() -> Path:
    return Path(WORKING_DIRECTORY) / WORKING_WORKBOOK_FILENAME


def workbook_exists(workbook_path: str | Path) -> bool:
    return Path(workbook_path).exists()


def ensure_working_workbook(input_workbook_path: str | Path) -> Path:
    working_path = resolve_working_workbook_path()
    working_path.parent.mkdir(parents=True, exist_ok=True)

    if not working_path.exists():
        copy2(input_workbook_path, working_path)

    return working_path


def reset_working_workbook(input_workbook_path: str | Path) -> Path:
    working_path = resolve_working_workbook_path()
    working_path.parent.mkdir(parents=True, exist_ok=True)
    copy2(input_workbook_path, working_path)
    return working_path


def _row_has_visible_value(row_values: list[object]) -> bool:
    return any(value is not None and str(value).strip() != "" for value in row_values)


def _build_unique_headers(header_values: list[object]) -> list[str]:
    headers: list[str] = []
    used_headers: dict[str, int] = {}

    for index, value in enumerate(header_values, start=1):
        base_header = str(value).strip() if value is not None and str(value).strip() else f"Column_{index}"
        duplicate_count = used_headers.get(base_header, 0)
        used_headers[base_header] = duplicate_count + 1
        header_name = base_header if duplicate_count == 0 else f"{base_header}_{duplicate_count + 1}"
        headers.append(header_name)

    return headers


def load_sheet_dataframe(
    workbook_path: str | Path,
    sheet_name: str,
    *,
    data_only: bool = True,
) -> tuple[pd.DataFrame, dict[str, Any]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=data_only)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet not found: {sheet_name}")

        worksheet = workbook[sheet_name]
        non_empty_rows: list[tuple[int, list[object]]] = []

        for row in worksheet.iter_rows(values_only=False):
            row_values = [cell.value for cell in row]
            if _row_has_visible_value(row_values):
                non_empty_rows.append((row[0].row, row_values))

        if not non_empty_rows:
            return pd.DataFrame(), {"header_excel_row": 1, "column_excel_indexes": {}}

        max_column_index = 0
        for _, row_values in non_empty_rows:
            for column_index, value in enumerate(row_values, start=1):
                if value is not None and str(value).strip() != "":
                    max_column_index = max(max_column_index, column_index)

        header_excel_row, header_values = non_empty_rows[0]
        header_values = header_values[:max_column_index]
        headers = _build_unique_headers(header_values)

        data_rows: list[list[object]] = []
        excel_row_numbers: list[int] = []

        for excel_row_number, row_values in non_empty_rows[1:]:
            trimmed_values = row_values[:max_column_index]
            padded_row = trimmed_values + [None] * (len(headers) - len(trimmed_values))
            data_rows.append(padded_row[: len(headers)])
            excel_row_numbers.append(excel_row_number)

        dataframe = pd.DataFrame(data_rows, columns=headers)
        dataframe["__excel_row_number"] = excel_row_numbers
        dataframe["__row_order"] = range(len(dataframe))
        dataframe["__row_key"] = dataframe["__excel_row_number"].astype(str)

        visible_columns = [column for column in dataframe.columns if not str(column).startswith("__")]
        dataframe = dataframe.dropna(how="all", subset=visible_columns).reset_index(drop=True)
        dataframe["__row_order"] = range(len(dataframe))
        dataframe["__row_key"] = dataframe["__excel_row_number"].astype(str)

        column_excel_indexes = {column_name: index for index, column_name in enumerate(headers, start=1)}
        metadata = {
            "sheet_name": sheet_name,
            "header_excel_row": header_excel_row,
            "column_excel_indexes": column_excel_indexes,
        }
        return dataframe, metadata
    finally:
        workbook.close()


def load_all_engine_data(workbook_path: str | Path, *, data_only: bool = True):
    master_dataframe, master_metadata = load_sheet_dataframe(workbook_path, MASTER_SHEET_NAME, data_only=data_only)
    bp_dataframe, bp_metadata = load_sheet_dataframe(workbook_path, BP_SHEET_NAME, data_only=data_only)
    reqd_dataframe, reqd_metadata = load_sheet_dataframe(workbook_path, REQD_MC_SHEET_NAME, data_only=data_only)

    return {
        "master": master_dataframe,
        "bp": bp_dataframe,
        "reqd_mc": reqd_dataframe,
        "metadata": {
            MASTER_SHEET_NAME: master_metadata,
            BP_SHEET_NAME: bp_metadata,
            REQD_MC_SHEET_NAME: reqd_metadata,
        },
    }


def _coerce_cell_value(value: object) -> object:
    if value is None or pd.isna(value):
        return None

    if isinstance(value, str):
        stripped_value = value.strip()
        if stripped_value == "":
            return None
        normalized_value = stripped_value.replace(",", "")
        try:
            numeric_value = float(normalized_value)
            return int(numeric_value) if numeric_value.is_integer() else numeric_value
        except ValueError:
            return stripped_value

    return value


def write_row_updates_to_workbook(
    workbook_path: str | Path,
    sheet_name: str,
    changed_values: dict[int, dict[str, object]],
    column_excel_indexes: dict[str, int],
) -> None:
    if not changed_values:
        return

    workbook = load_workbook(workbook_path, read_only=False, data_only=False)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet not found: {sheet_name}")

        worksheet = workbook[sheet_name]
        for excel_row_number, updates in changed_values.items():
            for column_name, value in updates.items():
                if column_name not in column_excel_indexes:
                    continue

                column_index = column_excel_indexes[column_name]
                target_cell = worksheet.cell(row=int(excel_row_number), column=int(column_index))
                existing_value = target_cell.value
                if isinstance(existing_value, str) and existing_value.startswith("="):
                    raise ValueError(
                        f"{sheet_name}!{target_cell.coordinate} is a formula cell. Please edit only input/driver cells."
                    )
                target_cell.value = _coerce_cell_value(value)

        workbook.save(workbook_path)
    finally:
        workbook.close()


def create_download_workbook_copy(working_workbook_path: str | Path) -> Path:
    output_directory = Path(OUTPUT_DIRECTORY)
    output_directory.mkdir(parents=True, exist_ok=True)
    output_path = output_directory / "Hyderabad_Manpower_Updated.xlsx"
    copy2(working_workbook_path, output_path)
    return output_path
