from __future__ import annotations

import re
from typing import BinaryIO

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from config.constants import (
    BP_READ_ONLY_COLUMNS,
    BP_SHEET_NAME,
    DRIVER_TABLE_HEIGHT,
    REQD_MC_EDITABLE_COLUMNS,
    REQD_MC_SHEET_NAME,
)
from core.excel_io import load_all_engine_data, write_row_updates_to_workbook
from core.excel_recalc import recalculate_workbook_with_excel
from core.formatting import build_formatted_display_dataframe
from core.totals import append_total_row, build_bp_kpis, build_reqd_mc_kpis
from ui.table_renderer import render_html_table
from ui.tab_cards import render_tab_cards


MONTH_PATTERN = re.compile(r"^[A-Z][a-z]{2}-\d{2}$")
EXCEL_DATE_HEADER_PATTERN = re.compile(r"^\d{4}-\d{2}-\d{2}(?: 00:00:00)?$")


def _visible_columns(dataframe: pd.DataFrame) -> list[str]:
    return [column_name for column_name in dataframe.columns if not str(column_name).startswith("__")]


def _detect_month_columns(dataframe: pd.DataFrame) -> list[str]:
    return [
        column_name
        for column_name in _visible_columns(dataframe)
        if MONTH_PATTERN.match(str(column_name)) or EXCEL_DATE_HEADER_PATTERN.match(str(column_name))
    ]


def _numeric_columns_present(dataframe: pd.DataFrame, candidates: list[str]) -> list[str]:
    return [column_name for column_name in candidates if column_name in dataframe.columns]


def _build_number_config(columns: list[str]):
    return {
        column_name: st.column_config.NumberColumn(
            label=column_name,
            format="%.2f",
            step=0.01,
        )
        for column_name in columns
    }


def _load_uploaded_bp_rows(uploaded_file: BinaryIO) -> list[list[object]]:
    uploaded_file.seek(0)
    workbook = load_workbook(uploaded_file, read_only=True, data_only=False)
    try:
        worksheet = workbook[BP_SHEET_NAME] if BP_SHEET_NAME in workbook.sheetnames else workbook.active
        rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
    finally:
        workbook.close()

    non_empty_rows = [
        row
        for row in rows
        if any(value is not None and str(value).strip() != "" for value in row)
    ]
    if not non_empty_rows:
        raise ValueError("Uploaded workbook does not contain any rows.")

    header_values = [str(value).strip() if value is not None else "" for value in non_empty_rows[0]]
    if len(header_values) < 2 or header_values[0] != "Section" or header_values[1] != "Machine":
        raise ValueError("Uploaded BP data must start with Section and Machine columns.")

    last_used_column = 0
    for row in non_empty_rows:
        for column_index, value in enumerate(row, start=1):
            if value is not None and str(value).strip() != "":
                last_used_column = max(last_used_column, column_index)

    return [row[:last_used_column] for row in non_empty_rows]


def _replace_bp_sheet_with_rows(rows: list[list[object]]) -> None:
    workbook = load_workbook(st.session_state.working_workbook_path, read_only=False, data_only=False)
    try:
        if BP_SHEET_NAME not in workbook.sheetnames:
            raise ValueError(f"Sheet not found: {BP_SHEET_NAME}")

        worksheet = workbook[BP_SHEET_NAME]
        max_row = max(worksheet.max_row, len(rows))
        max_column = max(worksheet.max_column, max(len(row) for row in rows))

        for row_index in range(1, max_row + 1):
            uploaded_row = rows[row_index - 1] if row_index <= len(rows) else []
            for column_index in range(1, max_column + 1):
                value = uploaded_row[column_index - 1] if column_index <= len(uploaded_row) else None
                worksheet.cell(row=row_index, column=column_index).value = value

        workbook.save(st.session_state.working_workbook_path)
    finally:
        workbook.close()


def _refresh_driver_data() -> None:
    engine_data = load_all_engine_data(st.session_state.working_workbook_path, data_only=True)
    st.session_state.master_dataframe = engine_data["master"]
    st.session_state.bp_dataframe = engine_data["bp"]
    st.session_state.reqd_mc_dataframe = engine_data["reqd_mc"]
    st.session_state.sheet_metadata = engine_data["metadata"]
    st.session_state.freeze_status = "Draft"


def _normalize_series(series: pd.Series) -> pd.Series:
    return series.fillna("").astype(str).str.replace(",", "", regex=False).str.strip()


def _extract_changed_values(
    original_dataframe: pd.DataFrame,
    edited_dataframe: pd.DataFrame,
    editable_columns: list[str],
) -> dict[int, dict[str, object]]:
    changed_updates: dict[int, dict[str, object]] = {}

    for column_name in editable_columns:
        if column_name not in edited_dataframe.columns:
            continue

        original_values = _normalize_series(original_dataframe[column_name])
        edited_values = _normalize_series(edited_dataframe[column_name])

        for excel_row_number, edited_value in edited_values.items():
            original_value = original_values.loc[excel_row_number]
            if edited_value == original_value:
                continue

            changed_updates.setdefault(int(excel_row_number), {})[column_name] = edited_dataframe.loc[
                excel_row_number,
                column_name,
            ]

    return changed_updates


def _apply_driver_updates(
    sheet_name: str,
    changed_updates: dict[int, dict[str, object]],
) -> None:
    metadata = st.session_state.sheet_metadata[sheet_name]
    write_row_updates_to_workbook(
        st.session_state.working_workbook_path,
        sheet_name,
        changed_updates,
        metadata["column_excel_indexes"],
    )

    recalculated, message = recalculate_workbook_with_excel(st.session_state.working_workbook_path)
    if recalculated:
        st.success("Changes applied and Excel formulas recalculated.")
    else:
        st.warning(f"Changes applied, but Excel automatic recalculation did not complete. {message}")

    _refresh_driver_data()
    st.rerun()


def _render_bp_upload_panel() -> None:
    st.markdown(
        """
        <div class="summary-banner">
            <div class="summary-banner-text">Upload BP data</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    upload_column, button_column = st.columns([5, 1.6])
    with upload_column:
        uploaded_file = st.file_uploader(
            "Upload BP Structured Excel file",
            type=["xlsx"],
            key="bp_structured_upload",
            label_visibility="collapsed",
        )

    with button_column:
        upload_clicked = st.button(
            "Upload Data",
            key="apply_bp_structured_upload",
            width="stretch",
            disabled=uploaded_file is None,
        )

    if upload_clicked and uploaded_file is not None:
        try:
            uploaded_rows = _load_uploaded_bp_rows(uploaded_file)
            _replace_bp_sheet_with_rows(uploaded_rows)
            recalculated, message = recalculate_workbook_with_excel(st.session_state.working_workbook_path)
            if recalculated:
                st.success("BP data uploaded and Excel formulas recalculated.")
            else:
                st.warning(f"BP data uploaded, but Excel automatic recalculation did not complete. {message}")
            _refresh_driver_data()
            st.rerun()
        except Exception as exc:
            st.error(f"BP upload failed: {exc}")


def render_bp_structured_tab() -> None:
    dataframe = st.session_state.bp_dataframe.copy()
    month_columns = _detect_month_columns(dataframe)

    render_tab_cards(build_bp_kpis(dataframe, month_columns))
    _render_bp_upload_panel()

    visible_dataframe = dataframe[_visible_columns(dataframe)].copy()
    display_dataframe = build_formatted_display_dataframe(visible_dataframe, numeric_columns=month_columns)
    render_html_table(display_dataframe, height=DRIVER_TABLE_HEIGHT, compact=False)

    st.markdown(
        """
        <div class="summary-banner">
            <div class="summary-banner-text">Edit BP monthly values</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    editor_dataframe = dataframe.copy().set_index("__excel_row_number", drop=True)
    editor_dataframe = editor_dataframe[_visible_columns(editor_dataframe)]

    editable_columns = month_columns
    disabled_columns = [column_name for column_name in editor_dataframe.columns if column_name not in editable_columns]

    edited_dataframe = st.data_editor(
        editor_dataframe,
        width="stretch",
        height=DRIVER_TABLE_HEIGHT,
        hide_index=True,
        disabled=disabled_columns,
        column_config=_build_number_config(editable_columns),
        key="editor_bp_structured",
    )

    changed_updates = _extract_changed_values(editor_dataframe, edited_dataframe, editable_columns)
    _, button_column, _ = st.columns([3.2, 1.6, 3.2])
    with button_column:
        if st.button(
            "Apply & Recalculate",
            key="apply_bp_structured",
            width="stretch",
            disabled=not bool(changed_updates),
        ):
            _apply_driver_updates(BP_SHEET_NAME, changed_updates)


def render_required_machines_tab() -> None:
    dataframe = st.session_state.reqd_mc_dataframe.copy()

    render_tab_cards(build_reqd_mc_kpis(dataframe))

    visible_dataframe = dataframe[_visible_columns(dataframe)].copy()
    display_dataframe = append_total_row(visible_dataframe, label_column="Section")
    display_dataframe = build_formatted_display_dataframe(
        display_dataframe,
        numeric_columns=[
            "Total Capacity/Day",
            "Total Capacity/Day/MC",
            "BP/Month",
            "BP/Day",
            "Available_MC",
            "Reqd_MC",
        ],
    )
    render_html_table(display_dataframe, height=DRIVER_TABLE_HEIGHT, compact=False)

    st.markdown(
        """
        <div class="summary-banner">
            <div class="summary-banner-text">Edit machine capacity drivers</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    editor_dataframe = dataframe.copy().set_index("__excel_row_number", drop=True)
    editor_dataframe = editor_dataframe[_visible_columns(editor_dataframe)]

    editable_columns = [column_name for column_name in REQD_MC_EDITABLE_COLUMNS if column_name in editor_dataframe.columns]
    disabled_columns = [column_name for column_name in editor_dataframe.columns if column_name not in editable_columns]

    edited_dataframe = st.data_editor(
        editor_dataframe,
        width="stretch",
        height=DRIVER_TABLE_HEIGHT,
        hide_index=True,
        disabled=disabled_columns,
        column_config=_build_number_config(editable_columns),
        key="editor_required_machines",
    )

    changed_updates = _extract_changed_values(editor_dataframe, edited_dataframe, editable_columns)
    _, button_column, _ = st.columns([3.2, 1.6, 3.2])
    with button_column:
        if st.button(
            "Apply & Recalculate",
            key="apply_required_machines",
            width="stretch",
            disabled=not bool(changed_updates),
        ):
            _apply_driver_updates(REQD_MC_SHEET_NAME, changed_updates)
